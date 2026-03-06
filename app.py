import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'

import streamlit as st
import cv2
import numpy as np
import pickle
import av
import threading
import pandas as pd
import serial
import serial.tools.list_ports
import time
from streamlit_webrtc import webrtc_streamer, VideoProcessorBase
from datetime import datetime, date
import tensorflow as tf
from face_utils import create_embedder, extract_face_crops, get_embedding
from train_model import retrain_system

# ─────────────────────────────────────────────
# Page Config
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Smart Attendance System",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    .stApp { background: #0f1117; }

    div[data-testid="metric-container"] {
        background: linear-gradient(135deg, #1e2130, #252a40);
        border: 1px solid #2d3250; border-radius: 14px;
        padding: 18px 22px; box-shadow: 0 4px 20px rgba(0,0,0,0.3);
    }
    div[data-testid="metric-container"] label {
        color: #8b92b3 !important; font-size: 0.78rem;
        letter-spacing: 0.08em; text-transform: uppercase;
    }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
        color: #e0e6ff !important; font-size: 2rem; font-weight: 700;
    }

    .status-card { border-radius: 14px; padding: 20px 24px; margin-bottom: 12px; border: 1px solid rgba(255,255,255,0.07); }
    .card-active   { background: linear-gradient(135deg, #0d2b1e, #143d28); border-color: #1a6640; }
    .card-inactive { background: linear-gradient(135deg, #2b1a0d, #3d2814); border-color: #6b4010; }
    .card-error    { background: linear-gradient(135deg, #2b0d0d, #3d1414); border-color: #8b2020; }

    .badge { display:inline-block; border-radius:20px; padding:3px 12px; font-size:0.73rem; font-weight:700; letter-spacing:0.05em; }
    .badge-active   { background:#1a6640; color:#5aedaa; }
    .badge-inactive { background:#6b4010; color:#ffaa55; }
    .badge-error    { background:#8b2020; color:#ff8080; }

    .card-title { font-size:1rem; font-weight:700; color:#e0e6ff; margin-bottom:8px; }
    .card-sub   { color:#9ba3c5; font-size:0.87rem; margin:4px 0; }
    .card-last  { color:#c5cbf0; font-size:0.92rem; margin-top:8px; }

    .section-header {
        font-size:1.1rem; font-weight:700; color:#c5cbf0;
        border-left:3px solid #5a6eff; padding-left:12px; margin:20px 0 12px 0;
    }
    .alert-info    { background:#0d1c2b; border:1px solid #1040a0; border-radius:10px; padding:12px 16px; color:#55aaff; margin:8px 0; font-size:0.9rem; }
    .alert-success { background:#0d2b1e; border:1px solid #1a6640; border-radius:10px; padding:12px 16px; color:#5aedaa; margin:8px 0; font-size:0.9rem; }
    .alert-warn    { background:#2b1f0d; border:1px solid #6b5010; border-radius:10px; padding:12px 16px; color:#ffcc55; margin:8px 0; font-size:0.9rem; }

    section[data-testid="stSidebar"] { background:#12151f !important; }
    section[data-testid="stSidebar"] .stButton>button {
        background: linear-gradient(135deg, #3d4fcc, #5a6eff) !important;
        color:white !important; border:none; border-radius:10px;
        width:100%; font-weight:600; padding:9px;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Constants & Paths
# ─────────────────────────────────────────────
MODEL_PATH  = os.path.join('models', 'attendance_model.h5')
LABEL_PATH  = os.path.join('models', 'label_map.pkl')
CSV_FILE    = "attendance.csv"
FP_MAP_FILE = "fp_map.pkl"

for d in ('data', 'models'):
    os.makedirs(d, exist_ok=True)

csv_lock = threading.Lock()

# ─────────────────────────────────────────────
# Session State Defaults
# ─────────────────────────────────────────────
DEFAULTS = {
    "arduino_serial": None,
    "arduino_port":   None,
    "fp_active":      False,
    "fp_status":      "Not started",
    "fp_last":        "—",
    "fp_thread":      None,
    "fp_stop":        False,
    "fp_map":         {},
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

if not st.session_state.fp_map and os.path.exists(FP_MAP_FILE):
    with open(FP_MAP_FILE, "rb") as f:
        st.session_state.fp_map = pickle.load(f)

# ─────────────────────────────────────────────
# Load AI Models
# ─────────────────────────────────────────────
@st.cache_resource
def load_models():
    embedder = create_embedder()
    try:
        model = tf.keras.models.load_model(MODEL_PATH)
        with open(LABEL_PATH, 'rb') as f:
            label_map = pickle.load(f)
    except:
        model, label_map = None, {}
    return embedder, model, label_map

embedder, model, label_map = load_models()

# ─────────────────────────────────────────────
# CSV Helper — shared by both systems
# ─────────────────────────────────────────────
def append_attendance(roll: str, name: str, method: str) -> bool:
    """Append one row. Returns True if written, False if duplicate for today."""
    today = str(date.today())
    with csv_lock:
        if os.path.exists(CSV_FILE):
            existing = pd.read_csv(CSV_FILE)
            dup = existing[
                (existing["Roll"].astype(str) == str(roll)) &
                (existing["Date"] == today) &
                (existing["Method"] == method)
            ]
            if not dup.empty:
                return False
        now = datetime.now()
        row = pd.DataFrame({
            "Date":   [today],
            "Time":   [now.strftime("%H:%M:%S")],
            "Roll":   [roll],
            "Name":   [name],
            "Method": [method],
            "Status": ["Present"],
        })
        if not os.path.exists(CSV_FILE):
            row.to_csv(CSV_FILE, index=False)
        else:
            row.to_csv(CSV_FILE, mode='a', header=False, index=False)
        return True

def load_csv() -> pd.DataFrame:
    if os.path.exists(CSV_FILE):
        return pd.read_csv(CSV_FILE)
    return pd.DataFrame(columns=["Date","Time","Roll","Name","Method","Status"])

# ─────────────────────────────────────────────
# WebRTC Processors
# ─────────────────────────────────────────────
class RegistrationProcessor(VideoProcessorBase):
    def __init__(self):
        self.captured_embeddings = []

    def recv(self, frame):
        img = frame.to_ndarray(format="bgr24")
        img = cv2.flip(img, 1)
        if len(self.captured_embeddings) < 20:
            faces = extract_face_crops(img)
            if faces:
                (x, y, w, h), face_crop = faces[0]
                cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)
                emb = get_embedding(face_crop, embedder)
                if emb is not None:
                    self.captured_embeddings.append(emb)
        cv2.putText(img, f"Captured: {len(self.captured_embeddings)}/20",
                    (30, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)
        return av.VideoFrame.from_ndarray(img, format="bgr24")


class AttendanceProcessor(VideoProcessorBase):
    def __init__(self):
        self.marked_session = set()

    def recv(self, frame):
        img = frame.to_ndarray(format="bgr24")
        img = cv2.flip(img, 1)
        if model is None:
            return av.VideoFrame.from_ndarray(img, format="bgr24")

        for (x, y, w, h), face_crop in extract_face_crops(img):
            emb = get_embedding(face_crop, embedder)
            if emb is None:
                continue

            pred     = model.predict(np.expand_dims(emb, axis=0), verbose=0)
            max_prob = float(np.max(pred))
            class_id = int(np.argmax(pred))

            if max_prob > 0.80:
                label = label_map.get(class_id, "Unknown")
                try:
                    roll, name = label.split('_', 1)
                except:
                    roll, name = "??", label
                color = (0, 255, 0)
                text  = f"{name} {int(max_prob*100)}%"
                if label not in self.marked_session:
                    if append_attendance(roll, name, "Face"):
                        print(f"✅ [Face] {name}")
                    self.marked_session.add(label)
            else:
                color = (0, 0, 255)
                text  = "Unknown"

            cv2.rectangle(img, (x, y), (x+w, y+h), color, 2)
            cv2.putText(img, text, (x, y-10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

        return av.VideoFrame.from_ndarray(img, format="bgr24")

# ─────────────────────────────────────────────
# Fingerprint Background Thread
# ─────────────────────────────────────────────
def fingerprint_loop():
    ser = st.session_state.arduino_serial
    if ser is None:
        st.session_state.fp_status = "❌ Arduino not connected"
        return

    cooldown     = {}
    COOLDOWN_SEC = 15

    while st.session_state.fp_active and not st.session_state.fp_stop:
        try:
            if ser.in_waiting:
                raw = ser.readline().decode("utf-8", errors="ignore").strip()

                if raw.startswith("ID:"):
                    fp_id = int(raw.split(":")[1])
                    label = st.session_state.fp_map.get(fp_id)
                    if label:
                        try:
                            roll, name = label.split('_', 1)
                        except:
                            roll, name = "??", label
                        st.session_state.fp_last   = f"{name} (ID {fp_id})"
                        st.session_state.fp_status = f"🟢 Matched: {name}"
                        now = time.time()
                        if cooldown.get(fp_id, 0) + COOLDOWN_SEC < now:
                            if append_attendance(roll, name, "Fingerprint"):
                                print(f"✅ [Fingerprint] {name}")
                            cooldown[fp_id] = now
                    else:
                        st.session_state.fp_status = f"⚠️ ID {fp_id} not mapped"
                        st.session_state.fp_last   = f"Unknown ID {fp_id}"

                elif raw == "FAIL":
                    st.session_state.fp_status = "❌ No match"
                elif raw == "PLACE":
                    st.session_state.fp_status = "👆 Waiting for finger…"
                elif raw == "READY":
                    st.session_state.fp_status = "🟡 Sensor ready"

        except Exception as e:
            st.session_state.fp_status = f"❌ Serial error: {e}"
            break
        time.sleep(0.08)

    st.session_state.fp_status = "🔴 Stopped"
    st.session_state.fp_active = False

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🎓 Smart Attendance")
    st.markdown("---")
    app_mode = st.radio("Navigate", [
        "📊 Dashboard",
        "📷 Mark Attendance (Face)",
        "👤 Register Student",
        "🔏 Fingerprint Setup",
        "📋 View Records",
    ], label_visibility="collapsed")

    st.markdown("---")
    st.markdown("### 🔌 Arduino Connection")

    ports    = [p.device for p in serial.tools.list_ports.comports()]
    sel_port = st.selectbox("COM Port", ["— none —"] + ports)
    baud     = st.selectbox("Baud Rate", [9600, 57600, 115200])

    cc, cd = st.columns(2)
    with cc:
        if st.button("Connect"):
            if sel_port != "— none —":
                try:
                    ser = serial.Serial(sel_port, baud, timeout=1)
                    st.session_state.arduino_serial = ser
                    st.session_state.arduino_port   = sel_port
                    st.success("Connected!")
                except Exception as e:
                    st.error(str(e))
    with cd:
        if st.button("Disconnect"):
            st.session_state.fp_stop   = True
            st.session_state.fp_active = False
            time.sleep(0.3)
            if st.session_state.arduino_serial:
                try:
                    st.session_state.arduino_serial.close()
                except:
                    pass
            st.session_state.arduino_serial = None
            st.session_state.arduino_port   = None
            st.rerun()

    if st.session_state.arduino_serial:
        st.markdown(
            f"<div style='color:#5aedaa;font-size:0.82rem;margin-bottom:6px'>● {st.session_state.arduino_port}</div>",
            unsafe_allow_html=True
        )
        if not st.session_state.fp_active:
            if st.button("▶ Start Fingerprint"):
                st.session_state.fp_active = True
                st.session_state.fp_stop   = False
                t = threading.Thread(target=fingerprint_loop, daemon=True)
                st.session_state.fp_thread = t
                t.start()
                st.rerun()
        else:
            if st.button("⏹ Stop Fingerprint"):
                st.session_state.fp_stop   = True
                st.session_state.fp_active = False
                st.rerun()

# ─────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────
if app_mode == "📊 Dashboard":
    st.title("📊 Attendance Dashboard")

    # ── Live Status Cards ────────────────────
    st.markdown("<div class='section-header'>Live System Status</div>", unsafe_allow_html=True)
    col_f, col_fp = st.columns(2)

    with col_f:
        face_ok  = model is not None
        f_cls    = "card-active" if face_ok else "card-error"
        f_badge  = "badge-active" if face_ok else "badge-error"
        f_label  = "MODEL READY" if face_ok else "NO MODEL"
        st.markdown(f"""
        <div class='status-card {f_cls}'>
            <div class='card-title'>📷 Face Recognition</div>
            <span class='badge {f_badge}'>{f_label}</span>
            <p class='card-sub' style='margin-top:10px'>
                {'AI model loaded — go to Mark Attendance' if face_ok
                 else 'Register at least one student to train the model'}
            </p>
            <p class='card-sub'>Students in model: <b style='color:#e0e6ff'>{len(label_map)}</b></p>
        </div>
        """, unsafe_allow_html=True)

    with col_fp:
        fp_on   = st.session_state.fp_active
        fp_conn = st.session_state.arduino_serial is not None
        fp_cls  = "card-active" if fp_on else ("card-inactive" if fp_conn else "card-error")
        fp_bdg  = "badge-active" if fp_on else ("badge-inactive" if fp_conn else "badge-error")
        fp_lbl  = "ACTIVE" if fp_on else ("CONNECTED" if fp_conn else "NO ARDUINO")
        st.markdown(f"""
        <div class='status-card {fp_cls}'>
            <div class='card-title'>🔏 Fingerprint Scanner</div>
            <span class='badge {fp_bdg}'>{fp_lbl}</span>
            <p class='card-sub' style='margin-top:10px'>{st.session_state.fp_status}</p>
            <p class='card-sub'>Port: <b style='color:#e0e6ff'>{st.session_state.arduino_port or '—'}</b></p>
            <p class='card-last'>Last matched: <b>{st.session_state.fp_last}</b></p>
        </div>
        """, unsafe_allow_html=True)

    # ── Today's Metrics ──────────────────────
    st.markdown("<div class='section-header'>Today's Summary</div>", unsafe_allow_html=True)
    df_all   = load_csv()
    today    = str(date.today())
    df_today = df_all[df_all["Date"] == today] if not df_all.empty else pd.DataFrame()

    total  = df_today["Name"].nunique()              if not df_today.empty else 0
    f_cnt  = len(df_today[df_today["Method"] == "Face"])          if not df_today.empty else 0
    fp_cnt = len(df_today[df_today["Method"] == "Fingerprint"])   if not df_today.empty else 0
    both   = int((df_today.groupby("Name")["Method"].nunique() > 1).sum()) if not df_today.empty else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Present",   total)
    m2.metric("Via Face",        f_cnt)
    m3.metric("Via Fingerprint", fp_cnt)
    m4.metric("Both Methods",    both)

    # ── Today's Table ────────────────────────
    st.markdown("<div class='section-header'>Today's Attendance Log</div>", unsafe_allow_html=True)
    if not df_today.empty:
        st.dataframe(
            df_today.sort_values("Time", ascending=False).reset_index(drop=True),
            use_container_width=True, hide_index=True
        )
    else:
        st.markdown("<div class='alert-info'>No attendance records yet for today.</div>",
                    unsafe_allow_html=True)

    if st.session_state.fp_active:
        time.sleep(3)
        st.rerun()

# ─────────────────────────────────────────────
# PAGE: MARK ATTENDANCE (FACE)
# ─────────────────────────────────────────────
elif app_mode == "📷 Mark Attendance (Face)":
    st.title("📷 Live Face Attendance")
    if model is None:
        st.markdown("<div class='alert-warn'>❌ No model found. Please register a student first.</div>",
                    unsafe_allow_html=True)
    else:
        st.markdown("<div class='alert-success'>✅ Model loaded. Show your face to mark attendance.</div>",
                    unsafe_allow_html=True)
        webrtc_streamer(
            key="attendance",
            video_processor_factory=AttendanceProcessor,
            rtc_configuration={"iceServers": [{"urls": ["stun:stun.l.google.com:19302"]}]},
            media_stream_constraints={"video": True, "audio": False},
        )

# ─────────────────────────────────────────────
# PAGE: REGISTER STUDENT
# ─────────────────────────────────────────────
elif app_mode == "👤 Register Student":
    st.title("👤 Register New Student")
    col1, col2 = st.columns(2)
    name = col1.text_input("Full Name")
    roll = col2.text_input("Roll Number")

    ctx = webrtc_streamer(
        key="registration",
        video_processor_factory=RegistrationProcessor,
        rtc_configuration={"iceServers": [{"urls": ["stun:stun.l.google.com:19302"]}]},
        media_stream_constraints={"video": True, "audio": False},
    )

    if st.button("💾 Save & Train"):
        if ctx.video_processor:
            captured = ctx.video_processor.captured_embeddings
            if len(captured) < 20:
                st.warning(f"Only {len(captured)}/20 frames captured. Keep your face in view!")
            elif not name or not roll:
                st.error("Please fill in both Name and Roll Number.")
            else:
                with open(f"data/{roll}_{name}.pkl", 'wb') as f:
                    pickle.dump(captured, f)
                st.success("✅ Data saved! Retraining model…")
                retrain_system()
                load_models.clear()
                st.success("🚀 Model updated! Switch to Mark Attendance.")

# ─────────────────────────────────────────────
# PAGE: FINGERPRINT SETUP
# ─────────────────────────────────────────────
elif app_mode == "🔏 Fingerprint Setup":
    st.title("🔏 Fingerprint Configuration")

    # Connection banner
    if st.session_state.arduino_serial:
        st.markdown(
            f"<div class='alert-success'>✅ Arduino connected on {st.session_state.arduino_port}</div>",
            unsafe_allow_html=True)
    else:
        st.markdown(
            "<div class='alert-warn'>⚠️ Arduino not connected. Select COM port in the sidebar.</div>",
            unsafe_allow_html=True)

    # Live scanner status card
    st.markdown("<div class='section-header'>Scanner Status</div>", unsafe_allow_html=True)
    fp_cls = "card-active" if st.session_state.fp_active else "card-inactive"
    fp_bdg = "badge-active" if st.session_state.fp_active else "badge-inactive"
    fp_lbl = "RUNNING" if st.session_state.fp_active else "STOPPED"
    st.markdown(f"""
    <div class='status-card {fp_cls}'>
        <div class='card-title'>🔏 Fingerprint System &nbsp;
            <span class='badge {fp_bdg}'>{fp_lbl}</span>
        </div>
        <p class='card-sub' style='margin-top:10px'>{st.session_state.fp_status}</p>
        <p class='card-last'>Last matched: <b>{st.session_state.fp_last}</b></p>
    </div>
    """, unsafe_allow_html=True)

    # Fingerprint ID → Student mapping
    st.markdown("<div class='section-header'>Sensor ID → Student Mapping</div>",
                unsafe_allow_html=True)
    st.markdown("""
    <div class='alert-info'>
    Each student's fingerprint is stored in a numbered slot on the sensor (ID 1, 2, 3…).
    Enrol fingers first using the Arduino sketch below, then map each ID to a student here.
    </div>
    """, unsafe_allow_html=True)

    with st.form("fp_map_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        fp_id   = c1.number_input("Sensor ID", min_value=1, step=1, value=1)
        fp_roll = c2.text_input("Roll Number")
        fp_name = c3.text_input("Student Name")
        if st.form_submit_button("➕ Add / Update Mapping"):
            if fp_roll and fp_name:
                st.session_state.fp_map[int(fp_id)] = f"{fp_roll}_{fp_name}"
                with open(FP_MAP_FILE, "wb") as f:
                    pickle.dump(st.session_state.fp_map, f)
                st.success(f"Mapped Sensor ID {int(fp_id)} → {fp_name} ({fp_roll})")
            else:
                st.error("Please fill in both Roll Number and Name.")

    if st.session_state.fp_map:
        rows = [
            {"Sensor ID": k, "Roll": v.split('_',1)[0], "Name": v.split('_',1)[-1]}
            for k, v in sorted(st.session_state.fp_map.items())
        ]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        del_opts = ["— select —"] + [
            f"ID {k} — {v.split('_',1)[-1]}"
            for k, v in sorted(st.session_state.fp_map.items())
        ]
        del_sel = st.selectbox("Remove a mapping", del_opts)
        if st.button("🗑️ Remove") and del_sel != "— select —":
            kid = int(del_sel.split()[1])
            del st.session_state.fp_map[kid]
            with open(FP_MAP_FILE, "wb") as f:
                pickle.dump(st.session_state.fp_map, f)
            st.rerun()
    else:
        st.markdown("<div class='alert-warn'>No mappings yet. Add one above.</div>",
                    unsafe_allow_html=True)

    # Arduino sketch embedded
    st.markdown("<div class='section-header'>Arduino Sketch (upload once to your board)</div>",
                unsafe_allow_html=True)
    st.code("""// Install via Library Manager: "Adafruit Fingerprint Sensor Library"

#include <Adafruit_Fingerprint.h>
#include <SoftwareSerial.h>

SoftwareSerial mySerial(2, 3);        // Sensor RX=pin2, TX=pin3
Adafruit_Fingerprint finger(&mySerial);

void setup() {
  Serial.begin(9600);                 // USB → PC baud (Streamlit reads this)
  finger.begin(57600);                // Sensor internal baud
  if (finger.verifyPassword()) {
    Serial.println("READY");
  } else {
    Serial.println("ERROR:SENSOR");
    while (true);
  }
}

void loop() {
  Serial.println("PLACE");           // Streamlit shows "Waiting for finger…"
  delay(800);

  if (finger.getImage()    != FINGERPRINT_OK) return;
  if (finger.image2Tz()    != FINGERPRINT_OK) return;
  if (finger.fingerSearch() == FINGERPRINT_OK) {
    Serial.print("ID:");
    Serial.println(finger.fingerID); // e.g.  "ID:3"
    delay(1500);
  } else {
    Serial.println("FAIL");
    delay(800);
  }
}

// ── ENROLMENT (run once per student, then re-flash main sketch) ──
// void enroll(uint8_t id) {
//   while (finger.getImage() != FINGERPRINT_OK);  finger.image2Tz(1);
//   while (finger.getImage() != FINGERPRINT_OK);  finger.image2Tz(2);
//   finger.createModel();
//   finger.storeModel(id);
//   Serial.print("Stored:"); Serial.println(id);
// }
""", language="cpp")

    # Wiring guide
    st.markdown("<div class='section-header'>Sensor Wiring (R307 / R503 / AS608)</div>",
                unsafe_allow_html=True)
    wiring = pd.DataFrame({
        "Sensor Pin": ["VCC", "GND", "TX", "RX"],
        "Arduino Pin": ["5 V", "GND", "Pin 2 (via 1 kΩ resistor)", "Pin 3"],
        "Notes": ["Power", "Ground", "Data to Arduino", "Data from Arduino"],
    })
    st.dataframe(wiring, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# PAGE: VIEW RECORDS
# ─────────────────────────────────────────────
elif app_mode == "📋 View Records":
    st.title("📋 Attendance Records")

    df_all = load_csv()

    if df_all.empty:
        st.markdown("<div class='alert-info'>No attendance records found yet.</div>",
                    unsafe_allow_html=True)
    else:
        # ── Filters ─────────────────────────
        st.markdown("<div class='section-header'>Filters</div>", unsafe_allow_html=True)
        f1, f2, f3 = st.columns(3)
        dates      = sorted(df_all["Date"].unique(), reverse=True)
        methods    = ["All"] + sorted(df_all["Method"].unique().tolist())
        names      = ["All"] + sorted(df_all["Name"].unique().tolist())

        sel_date   = f1.selectbox("Date",    ["All"] + dates)
        sel_method = f2.selectbox("Method",  methods)
        sel_name   = f3.selectbox("Student", names)

        df_f = df_all.copy()
        if sel_date   != "All": df_f = df_f[df_f["Date"]   == sel_date]
        if sel_method != "All": df_f = df_f[df_f["Method"] == sel_method]
        if sel_name   != "All": df_f = df_f[df_f["Name"]   == sel_name]
        df_f = df_f.sort_values(["Date","Time"], ascending=[False,False]).reset_index(drop=True)

        # ── Summary metrics ──────────────────
        st.markdown("<div class='section-header'>Summary</div>", unsafe_allow_html=True)
        m1, m2, m3 = st.columns(3)
        m1.metric("Total Records",   len(df_f))
        m2.metric("Unique Students", df_f["Name"].nunique())
        m3.metric("Days Covered",    df_f["Date"].nunique())

        # ── Table ────────────────────────────
        st.markdown("<div class='section-header'>Records</div>", unsafe_allow_html=True)
        st.dataframe(df_f, use_container_width=True, hide_index=True)

        # ── Export ───────────────────────────
        st.markdown("<div class='section-header'>Export</div>", unsafe_allow_html=True)
        ec1, ec2 = st.columns(2)

        fname_base = sel_date if sel_date != "All" else "all"

        with ec1:
            st.download_button(
                label="⬇️ Download CSV",
                data=df_f.to_csv(index=False).encode(),
                file_name=f"attendance_{fname_base}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with ec2:
            try:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                buf = io.BytesIO()
                wb  = Workbook()
                ws  = wb.active
                ws.title = "Attendance"

                thin   = Side(style="thin", color="D0D0D0")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                h_fill = PatternFill("solid", fgColor="3D4FCC")
                h_font = Font(bold=True, color="FFFFFF", size=11)
                g_fill = PatternFill("solid", fgColor="D6F5E3")   # Face rows
                b_fill = PatternFill("solid", fgColor="D6E8F5")   # Fingerprint rows

                for ci, col_name in enumerate(df_f.columns, 1):
                    cell = ws.cell(row=1, column=ci, value=col_name)
                    cell.fill = h_fill; cell.font = h_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                    ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col_name)+4)

                for ri, row in enumerate(df_f.itertuples(index=False), 2):
                    row_fill = g_fill if getattr(row, "Method", "") == "Face" else b_fill
                    for ci, val in enumerate(row, 1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        cell.fill = row_fill
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border

                wb.save(buf); buf.seek(0)
                st.download_button(
                    label="⬇️ Download Excel",
                    data=buf,
                    file_name=f"attendance_{fname_base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except ImportError:
                st.info("Run `pip install openpyxl` to enable Excel export.")

        if st.button("🔄 Refresh"):
            st.rerun()